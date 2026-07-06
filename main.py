from flask import Flask, request, jsonify
import openpyxl
from openpyxl.styles import PatternFill
import urllib.request
import json, io, base64, os, re, traceback

app = Flask(__name__)

registry_state = {}

RED_FILL    = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

GROQ_API_KEY = os.environ.get('GROQ_API_KEY', '')

# Резервный список ключевых слов (если Groq недоступен)
PROHIBITED_STEMS = [
    'мазь', 'таблет', 'лекарств', 'бижутер', 'украшен',
    'ожерель', 'колье', 'чокер', 'кулон',
    'серьг', 'сережк', 'кафф', 'клипс', 'пусет',
    'браслет', 'оружи', 'брошь', 'подвеск',
    'золот', 'серебр', 'цепочк', 'кольц',
]
EXCLUSION_WORDS = [
    'посудомоеч', 'омывател', 'стиральн', 'наушник',
    'зажим', 'уплотн', 'держател', 'люверс',
    'конфет', 'миск', 'салатник', 'пуговиц',
    'маска', 'для лица', 'тени для', 'мюли', 'туфл', 'сланц', 'кроссов', 'сабо',
    'нержавеющ', 'из стали', 'наклейк', 'пикгард', 'гитар',
    'скраб', 'крем', 'бейдж', 'косметик',
    'клатч', 'пришивн', 'велосип', 'мотоцикл', 'ступиц', 'зубьев', 'концепц',
    'качел', 'гамак', 'пустые', 'ошейник', 'для кошек', 'для собак',
]


def is_prohibited_keyword(product):
    lower = product.lower()
    if any(excl in lower for excl in EXCLUSION_WORDS):
        return False
    return any(stem in lower for stem in PROHIBITED_STEMS)


GROQ_PROMPT_TEMPLATE = """Ты таможенный инспектор Таджикистана. Из списка товаров определи ТОЛЬКО запрещённые к ввозу.

ЗАПРЕЩЁННЫЕ (помечай):
- Ювелирные украшения: серьги, кольца ювелирные/обручальные/помолвочные, браслеты, цепочки на шею, кулоны, подвески на шею, броши, колье, ожерелья, чокеры, пусеты, каффы, бижутерия
- Лекарства и мази: таблетки лечебные, мази медицинские, крема лечебные от болезней (псориаз, грибок, артрит, геморрой и т.д.)

РАЗРЕШЁННЫЕ (НЕ помечай ни при каких условиях):
- Часы любые
- Чехлы для телефонов, электроника, аксессуары
- Обувь (кроссовки, сабо, туфли, сланцы и т.д.)
- Косметика (крема для лица/ног уходовые, маски для волос, масла для волос, скрабы, тени, помады, духи, солнцезащитные крема)
- Сумки, рюкзаки, клатчи (даже на цепочке)
- Посуда
- Зоотовары (ошейники, пасты для кошек, поводки)
- Товары для ногтей (верхние формы, фольга, органайзеры)
- Игрушки, антистресс
- Книги
- Рыболовные снасти
- Строительные материалы
- Брелоки
- Кольца для волос, украшения для кос/дредов
- Кольца уплотнительные, зажимы, держатели (промышленные)
- Таблетки для посудомойки/бассейна, таблетницы
- Тесты на беременность, глюкометры (медицинские устройства, не лекарства)
- Пистолеты для пирсинга
- Этикетки, наклейки
- Затирки, клей, герметики
- Товары-заменители, рандомные товары

Товары (индекс. название):
{numbered}

Ответь ТОЛЬКО JSON массивом индексов запрещённых товаров. Только цифры через запятую в скобках [].
Пример ответа: [2, 5, 11]
Если запрещённых нет: []"""


def call_groq_chunk(chunk_products, offset):
    """Отправляет один чанк товаров в Groq, возвращает set глобальных индексов запрещённых."""
    numbered = '\n'.join(f"{offset + i}. {p}" for i, p in enumerate(chunk_products))
    prompt = GROQ_PROMPT_TEMPLATE.format(numbered=numbered)

    data = json.dumps({
        "model": "llama-3.3-70b-versatile",
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0,
        "max_tokens": 2000,
    }).encode('utf-8')

    req = urllib.request.Request(
        'https://api.groq.com/openai/v1/chat/completions',
        data=data,
        headers={
            'Authorization': f'Bearer {GROQ_API_KEY}',
            'Content-Type': 'application/json',
        }
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        result = json.loads(resp.read())

    content = result['choices'][0]['message']['content'].strip()
    print(f"Groq chunk offset={offset} response: {content[:200]}", flush=True)

    match = re.search(r'\[[\d,\s]*\]', content)
    if match:
        return set(json.loads(match.group()))
    return set()


def classify_products_groq(products):
    """Отправляет список уникальных товаров в Groq чанками, возвращает set индексов запрещённых."""
    if not products or not GROQ_API_KEY:
        return None  # None = использовать fallback

    CHUNK_SIZE = 400
    prohibited_indices = set()

    try:
        for offset in range(0, len(products), CHUNK_SIZE):
            chunk = products[offset:offset + CHUNK_SIZE]
            chunk_result = call_groq_chunk(chunk, offset)
            prohibited_indices |= chunk_result
            if offset + CHUNK_SIZE < len(products):
                import time; time.sleep(2)  # пауза между чанками

        return prohibited_indices

    except Exception as e:
        print(f"Groq error: {e}", flush=True)
        return None  # None = fallback на ключевые слова


def get_usd_tjs_rate():
    url = 'https://open.er-api.com/v6/latest/USD'
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req, timeout=10) as response:
        data = json.loads(response.read())
    return data['rates']['TJS']


def find_col(headers, *names):
    for name in names:
        for i, h in enumerate(headers):
            if h and str(h).strip().upper() == name.strip().upper():
                return i
    for name in names:
        for i, h in enumerate(headers):
            if h and name.lower() in str(h).lower():
                return i
    return None


@app.route('/')
@app.route('/health')
def health():
    return jsonify({'status': 'ok'})


@app.route('/state', methods=['POST'])
def save_state():
    global registry_state
    registry_state = request.get_json()
    return jsonify({'ok': True})


@app.route('/state', methods=['GET'])
def get_state():
    return jsonify(registry_state)


@app.route('/process', methods=['POST'])
def process():
    try:
        # Поддержка двух форматов: JSON (n8n) и form-data
        if request.is_json:
            payload = request.get_json()
            usd_rate_param = payload.get('usdRate')
            file_bytes = base64.b64decode(payload['fileBase64'])
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        else:
            usd_rate_param = request.form.get('usdRate')
            file_bytes_io = request.files['file']
            wb = openpyxl.load_workbook(file_bytes_io)

        rate = float(usd_rate_param) if usd_rate_param else get_usd_tjs_rate()
        limit = 200 * rate

        ws = wb.active

        headers = [cell.value for cell in ws[1]]

        sum_col     = find_col(headers, 'Сумма')
        name_col    = find_col(headers, 'Имя получателя')
        surname_col = find_col(headers, 'Фамилия получателя')
        phone_col   = find_col(headers, 'Контактный номер', 'Телефон')
        product_col = find_col(headers, 'Наименование товара')

        print(f"Cols: sum={sum_col} name={name_col} surname={surname_col} phone={phone_col} product={product_col}", flush=True)

        max_data_col = max(
            (i + 1 for i, h in enumerate(headers) if h is not None),
            default=1
        )

        # Первый проход: собрать данные и суммы
        groups = {}
        rows_data = []

        for row in ws.iter_rows(min_row=2):
            name    = str(row[name_col].value    or '') if name_col    is not None else ''
            surname = str(row[surname_col].value  or '') if surname_col is not None else ''
            phone   = str(row[phone_col].value    or '') if phone_col   is not None else ''
            product = str(row[product_col].value  or '') if product_col is not None else ''
            full_name = f"{name} {surname}".strip()
            key = full_name + phone

            try:
                amount = float(row[sum_col].value or 0) if sum_col is not None else 0.0
            except (ValueError, TypeError):
                amount = 0.0

            groups[key] = groups.get(key, 0) + amount
            rows_data.append({
                'row': row, 'key': key,
                'full_name': full_name, 'product': product
            })

        # Классификация товаров через Groq (один батч-запрос)
        unique_products = list(dict.fromkeys(
            item['product'] for item in rows_data if item['product'].strip()
        ))
        prohibited_set = classify_products_groq(unique_products)
        use_groq = prohibited_set is not None

        if use_groq:
            prohibited_products = {unique_products[i] for i in prohibited_set}
            print(f"Groq classified {len(prohibited_products)} prohibited out of {len(unique_products)}", flush=True)
        else:
            print("Groq unavailable, using keyword fallback", flush=True)

        # Второй проход: красить и собирать нарушителей
        violators, prohibited_items = [], []
        seen_violators, seen_prohibited = set(), set()

        for item in rows_data:
            row     = item['row']
            key     = item['key']
            name    = item['full_name']
            product = item['product']

            if use_groq:
                row_prohibited = product in prohibited_products
            else:
                row_prohibited = is_prohibited_keyword(product)

            row_violator = groups[key] > limit

            # Считаем нарушителей со ВСЕХ строк (включая уже покрашенные)
            if row_violator:
                if name not in seen_violators:
                    seen_violators.add(name)
                    violators.append({
                        'name': name,
                        'amountUSD': round(groups[key] / rate, 2)
                    })
            elif row_prohibited:
                if name + product not in seen_prohibited:
                    seen_prohibited.add(name + product)
                    prohibited_items.append({
                        'name': name,
                        'product': product,
                        'amountUSD': round(groups[key] / rate, 2)
                    })

            # Не перекрашивать уже покрашенные строки
            try:
                existing = row[0].fill
                if existing and existing.fill_type not in (None, 'none') and \
                   existing.fgColor.rgb not in ('00000000', 'FF000000'):
                    continue
            except Exception:
                pass

            if row_violator:
                for cell in row[:max_data_col]:
                    cell.fill = YELLOW_FILL
            elif row_prohibited:
                for cell in row[:max_data_col]:
                    cell.fill = RED_FILL

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return jsonify({
            'success': True,
            'violators': violators,
            'prohibited': prohibited_items,
            'totalProcessed': ws.max_row - 1,
            'usdRate': rate,
            'usedGroq': use_groq,
            'fileBase64': base64.b64encode(output.read()).decode(),
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'trace': traceback.format_exc()
        }), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port)
